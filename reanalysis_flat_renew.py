# -*- coding: utf-8 -*-
"""
@author: Danil Borchevkin

"""

from openpyxl import load_workbook
import csv


def drange_up(start, stop, step):
    ret = []
    r = start
    while r <= stop:
        ret.append(r)
        r += step
    return ret

def drange_down(start, stop, step):
    ret = []
    r = start
    while r >= stop:
        ret.append(r)
        r += step
    return ret

def get_indicies_of_values(list_of_values, start_value, end_value, shift=0):
    indicies_list = []

    for i,value in enumerate(list_of_values):
        if start_value <= value <= end_value:
            indicies_list.append(i + shift)

    return indicies_list

def parse__reanalysis_file(path, filename, start_lat, end_lat, start_long, end_long):
    # Because reading of lats and longs very high load task we define it
    LAT_LIST = drange_down(65, 35, 0.25)
    LONG_LIST = drange_up(0, 40, 0.25)
    SHIFT = 2

    # TODO change workdir
    workbook = load_workbook(filename=path+filename, read_only=True)
    sheets = workbook.sheetnames
    
    for sheet in sheets:
        print('Work with sheet [' + sheet + ']')
        
        # 0. Activate sheet
        worksheet = workbook[sheet]

        # 1. Filter lats and longs according to params and save it's indexes inside sheet
        desired_lats_index = get_indicies_of_values(LAT_LIST, end_lat, start_lat)
        desired_longs_index = get_indicies_of_values(LONG_LIST, start_long, end_long)
        
        # 2. Create output list
        # ws.cell method very very slow. 
        # sheet[ row ][ col ].value slow too. 
        # so we should use iter_rows 
        # iter_columns not available for ReadOnly workbooks
        output_list = []
        
        # Create 'borders' of desired range
        min_row = desired_lats_index[0]+SHIFT
        max_row = desired_lats_index[-1]+SHIFT
        min_col = desired_longs_index[0]+SHIFT
        max_col = desired_longs_index[-1]+SHIFT

        for lat_i,row in enumerate(worksheet.iter_rows(
                                        min_row=min_row,
                                        max_row=max_row,
                                        min_col=min_col,
                                        max_col=max_col,
                                        values_only=True)):
            for long_i,value in enumerate(row):
                output_list.append([
                    LONG_LIST[long_i+min_col-SHIFT],  #long
                    LAT_LIST[lat_i+min_row-SHIFT],    #lat
                    value               #value
                ])

        # 3. Write values to CSV
        with open('./output/' + filename + '_' + sheet + '.txt', 'w', newline='') as f:
            writer = csv.writer(f, delimiter='\t')
            for line in output_list:
                writer.writerow(line)

if __name__ == "__main__":
    print('Script is started')

    # 1. Change start and end values here
    # WARNING! Please see source file
    # - lats starts from 90 and ends by -90
    # - longs starts from -180 and ends by 179.375
    START_LAT = 65
    END_LAT = 35
    START_LONG = 0
    END_LONG = 40

    # 2. CHANGE FILENAME HERE. FILE MUST TO BE IN THE SAME FOLDER
    parse__reanalysis_file('./input/', 'V wind 975 hPa march 2020.xlsx', START_LAT, END_LAT, START_LONG, END_LONG) 

    # 3. Enjoy your files in the same directory =))

    print('script is ended')