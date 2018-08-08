# -*- coding: utf-8 -*-
"""
Created on Wed Aug  8 22:26:15 2018

@author: Danil Borchevkin

Dimension of the data 
1-42 by lat (rows)
1-50 by long (columns)

"""

from openpyxl import load_workbook
import csv

LATS_DIM = 42
LONG_DIM = 50

def parse_file(path, filename):
    # TODO change workdir
    workbook = load_workbook(filename=filename, read_only=True)
    
    sheets = workbook.sheetnames
    
    for sheet in sheets:
        print('Work with sheet [' + sheet + ']')
        
        #Activate sheet
        worksheet = workbook[sheet]
        
        #Read lats to var
        lats = []
        for i in range(2, LATS_DIM + 1):
            lats.append(worksheet.cell(row=i, column=1).value)
                    
        # Iterate over longs writing longs and wind to CSV
        with open(filename + '_' + sheet + '.csv', 'w', newline='') as f:
            writer = csv.writer(f, delimiter='\t')
            
            for i in range(2, LONG_DIM + 1):
                for j in range(len(lats)):
                    writer.writerow(
                            [lats[j],                                   #lat
                             worksheet.cell(row=1, column=i).value,     #long
                             worksheet.cell(row=j+2, column=i).value    #wind
                             ])

print('Script is started')
parse_file('', 'test.xlsx') # CHANGE FILENAME HERE. FILE MUST TO BE IN THE SAME FOLDER
print('sreipt is ended')